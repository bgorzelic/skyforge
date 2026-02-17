"""Reporter — export analysis, segment, and detection data to CSV and Excel formats."""

from __future__ import annotations

import csv
import json
from datetime import datetime
from pathlib import Path


def export_analysis_csv(analysis_dir: Path, output: Path) -> Path:
    """Export frame-level analysis data to a flat CSV file.

    Walks analysis_dir for ``*/analysis.json`` files (VideoAnalysis dicts)
    and flattens each ``frame_analyses`` entry into a CSV row.

    Args:
        analysis_dir: Directory containing per-video analysis subdirectories.
        output: Destination CSV path.

    Returns:
        The output path written.
    """
    fieldnames = [
        "source",
        "timestamp",
        "blur_score",
        "brightness",
        "contrast",
        "motion_score",
        "is_dark",
        "is_overexposed",
        "is_blurry",
    ]

    with output.open("w", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()

        for analysis_json in sorted(analysis_dir.rglob("*/analysis.json")):
            data = json.loads(analysis_json.read_text())
            source = Path(data.get("source_file", analysis_json.parent.name)).stem

            for frame in data.get("frame_analyses", []):
                writer.writerow(
                    {
                        "source": source,
                        "timestamp": frame.get("timestamp", 0.0),
                        "blur_score": frame.get("blur_score", 0.0),
                        "brightness": frame.get("brightness", 0.0),
                        "contrast": frame.get("contrast", 0.0),
                        "motion_score": frame.get("motion_score", 0.0),
                        "is_dark": frame.get("is_dark", False),
                        "is_overexposed": frame.get("is_overexposed", False),
                        "is_blurry": frame.get("is_blurry", False),
                    }
                )

    return output


def export_segments_csv(analysis_dir: Path, output: Path) -> Path:
    """Export segment selection data to a CSV file.

    Walks analysis_dir for ``selects_*.json`` files and writes each
    segment entry as a CSV row.

    Args:
        analysis_dir: Directory containing selects JSON files.
        output: Destination CSV path.

    Returns:
        The output path written.
    """
    fieldnames = [
        "source_file",
        "segment_id",
        "start_time",
        "end_time",
        "duration",
        "confidence",
        "reason_tags",
    ]

    with output.open("w", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()

        for selects_json in sorted(analysis_dir.glob("selects_*.json")):
            data = json.loads(selects_json.read_text())

            for segment in data.get("segments", []):
                writer.writerow(
                    {
                        "source_file": segment.get("source_file", ""),
                        "segment_id": segment.get("segment_id", 0),
                        "start_time": segment.get("start_time", 0.0),
                        "end_time": segment.get("end_time", 0.0),
                        "duration": segment.get("duration", 0.0),
                        "confidence": segment.get("confidence", 0.0),
                        "reason_tags": ";".join(segment.get("reason_tags", [])),
                    }
                )

    return output


def export_detections_csv(detections_dir: Path, output: Path) -> Path:
    """Export object detection data to a CSV file.

    Walks detections_dir for ``*_detections.json`` files and writes one
    row per frame, summarising the detections found.

    Args:
        detections_dir: Directory containing detection JSON files.
        output: Destination CSV path.

    Returns:
        The output path written.
    """
    fieldnames = [
        "source",
        "frame_idx",
        "timestamp_s",
        "total_detections",
        "classes",
    ]

    with output.open("w", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()

        for det_json in sorted(detections_dir.glob("*_detections.json")):
            data = json.loads(det_json.read_text())
            source = det_json.stem.removesuffix("_detections")

            for frame in data.get("frames", []):
                detections = frame.get("detections", [])
                unique_classes = sorted({d.get("class_name", "unknown") for d in detections})
                writer.writerow(
                    {
                        "source": source,
                        "frame_idx": frame.get("frame_idx", 0),
                        "timestamp_s": frame.get("timestamp_s", 0.0),
                        "total_detections": len(detections),
                        "classes": ";".join(unique_classes),
                    }
                )

    return output


def export_project_excel(project_dir: Path, output: Path) -> Path:
    """Export a multi-sheet Excel workbook summarising the project.

    Requires openpyxl. If not installed, raises ``ImportError`` with an
    install hint.

    Sheets created:
        - **Summary** -- project name, date, file counts.
        - **Frames** -- same data as ``export_analysis_csv``.
        - **Segments** -- same data as ``export_segments_csv``.
        - **Detections** -- same data as ``export_detections_csv`` (only if
          ``07_DETECTIONS`` exists).

    Args:
        project_dir: Root of the flight project.
        output: Destination ``.xlsx`` path.

    Returns:
        The output path written.

    Raises:
        ImportError: If openpyxl is not available.
    """
    try:
        from openpyxl import Workbook
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel export. Install with: pip install 'skyforge[reports]'"
        ) from None

    wb = Workbook()

    # ------------------------------------------------------------------ Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"  # type: ignore[union-attr]

    project_meta = _load_project_meta(project_dir)
    analysis_dir = project_dir / "03_ANALYSIS"
    detections_dir = project_dir / "07_DETECTIONS"

    analysis_files = list(analysis_dir.rglob("*/analysis.json")) if analysis_dir.exists() else []
    selects_files = list(analysis_dir.glob("selects_*.json")) if analysis_dir.exists() else []

    summary_rows: list[tuple[str, str]] = [
        ("Project Name", project_meta.get("name", project_dir.name)),
        ("Report Date", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Status", project_meta.get("status", "unknown")),
        ("Analysis Files", str(len(analysis_files))),
        ("Selects Files", str(len(selects_files))),
    ]
    if detections_dir.exists():
        det_count = len(list(detections_dir.glob("*_detections.json")))
        summary_rows.append(("Detection Files", str(det_count)))

    for row in summary_rows:
        ws_summary.append(row)  # type: ignore[union-attr]

    _set_column_widths(ws_summary, [25, 40])  # type: ignore[arg-type]

    # ------------------------------------------------------------------ Frames
    ws_frames = wb.create_sheet("Frames")
    frame_headers = [
        "source",
        "timestamp",
        "blur_score",
        "brightness",
        "contrast",
        "motion_score",
        "is_dark",
        "is_overexposed",
        "is_blurry",
    ]
    ws_frames.append(frame_headers)

    for analysis_json in sorted(analysis_files):
        data = json.loads(analysis_json.read_text())
        source = Path(data.get("source_file", analysis_json.parent.name)).stem
        for frame in data.get("frame_analyses", []):
            ws_frames.append(
                [
                    source,
                    frame.get("timestamp", 0.0),
                    frame.get("blur_score", 0.0),
                    frame.get("brightness", 0.0),
                    frame.get("contrast", 0.0),
                    frame.get("motion_score", 0.0),
                    frame.get("is_dark", False),
                    frame.get("is_overexposed", False),
                    frame.get("is_blurry", False),
                ]
            )

    _set_column_widths(ws_frames, [25, 12, 12, 12, 12, 12, 10, 14, 10])

    # ---------------------------------------------------------------- Segments
    ws_segments = wb.create_sheet("Segments")
    seg_headers = [
        "source_file",
        "segment_id",
        "start_time",
        "end_time",
        "duration",
        "confidence",
        "reason_tags",
    ]
    ws_segments.append(seg_headers)

    for selects_json in sorted(selects_files):
        data = json.loads(selects_json.read_text())
        for segment in data.get("segments", []):
            ws_segments.append(
                [
                    segment.get("source_file", ""),
                    segment.get("segment_id", 0),
                    segment.get("start_time", 0.0),
                    segment.get("end_time", 0.0),
                    segment.get("duration", 0.0),
                    segment.get("confidence", 0.0),
                    ";".join(segment.get("reason_tags", [])),
                ]
            )

    _set_column_widths(ws_segments, [30, 12, 12, 12, 12, 12, 40])

    # -------------------------------------------------------------- Detections
    if detections_dir.exists():
        ws_det = wb.create_sheet("Detections")
        det_headers = ["source", "frame_idx", "timestamp_s", "total_detections", "classes"]
        ws_det.append(det_headers)

        for det_json in sorted(detections_dir.glob("*_detections.json")):
            data = json.loads(det_json.read_text())
            source = det_json.stem.removesuffix("_detections")
            for frame in data.get("frames", []):
                detections = frame.get("detections", [])
                unique_classes = sorted({d.get("class_name", "unknown") for d in detections})
                ws_det.append(
                    [
                        source,
                        frame.get("frame_idx", 0),
                        frame.get("timestamp_s", 0.0),
                        len(detections),
                        ";".join(unique_classes),
                    ]
                )

        _set_column_widths(ws_det, [25, 12, 12, 18, 40])

    wb.save(output)
    return output


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _load_project_meta(project_dir: Path) -> dict:
    """Load project.json metadata, falling back to directory name."""
    meta_path = project_dir / "project.json"
    if meta_path.exists():
        return json.loads(meta_path.read_text())
    return {"name": project_dir.name, "status": "unknown"}


def _set_column_widths(ws, widths: list[int]) -> None:
    """Set column widths on a worksheet by index."""
    from openpyxl.utils import get_column_letter

    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
